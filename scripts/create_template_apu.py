"""
Script para crear el Template_APUS.xlsx con estructura real
Basado en la estructura de APUs con 4 categorías separadas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "TEMPLATE"

    # Estilos
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Fill colors
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_category = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_subtotal = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    current_row = 1

    # ===== INFORMACIÓN DEL RUBRO =====
    ws[f'A{current_row}'] = "CODIGO:"
    ws[f'A{current_row}'].font = header_font
    ws[f'B{current_row}'] = ""  # Se llenará con el código del rubro
    current_row += 1

    ws[f'A{current_row}'] = "DESCRIPCION:"
    ws[f'A{current_row}'].font = header_font
    ws[f'B{current_row}'] = ""  # Se llenará con la descripción
    current_row += 1

    ws[f'A{current_row}'] = "UNIDAD:"
    ws[f'A{current_row}'].font = header_font
    ws[f'B{current_row}'] = ""  # Se llenará con la unidad
    current_row += 2  # Espacio

    # ===== SECCIÓN EQUIPOS =====
    ws[f'A{current_row}'] = "EQUIPOS"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].fill = fill_category
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 1

    # Headers EQUIPOS
    headers_equipos = ['DESCRIPCION', 'CANTIDAD', 'TARIFA', 'COSTO HORA', 'RENDIMIENTO', 'COSTO']
    subheaders_equipos = ['', 'A', 'B', 'C = A x B', 'R', 'D = C x R']

    for col_idx, (header, subheader) in enumerate(zip(headers_equipos, subheaders_equipos), start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    for col_idx, subheader in enumerate(subheaders_equipos, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=subheader)
        cell.font = Font(italic=True, size=9)
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    # Filas de ejemplo para EQUIPOS (3 filas)
    equipos_start_row = current_row
    for i in range(3):
        for col_idx in range(1, 7):
            cell = ws.cell(row=current_row, column=col_idx, value="")
            cell.border = border_thin
        current_row += 1

    # Subtotal EQUIPOS
    ws[f'A{current_row}'] = "SUBTOTAL M"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = fill_subtotal
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'F{current_row}'].border = border_thin
    ws[f'F{current_row}'].fill = fill_subtotal
    current_row += 2  # Espacio

    # ===== SECCIÓN MANO DE OBRA =====
    ws[f'A{current_row}'] = "MANO DE OBRA"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].fill = fill_category
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 1

    # Headers MANO DE OBRA
    headers_mano = ['DESCRIPCION', 'CANTIDAD', 'JORNAL /HR', 'COSTO HORA', 'RENDIMIENTO', 'COSTO']
    subheaders_mano = ['', 'A', 'B', 'C = A x B', 'R', 'D = C x R']

    for col_idx, (header, subheader) in enumerate(zip(headers_mano, subheaders_mano), start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    for col_idx, subheader in enumerate(subheaders_mano, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=subheader)
        cell.font = Font(italic=True, size=9)
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    # Filas de ejemplo para MANO DE OBRA (3 filas)
    mano_start_row = current_row
    for i in range(3):
        for col_idx in range(1, 7):
            cell = ws.cell(row=current_row, column=col_idx, value="")
            cell.border = border_thin
        current_row += 1

    # Subtotal MANO DE OBRA
    ws[f'A{current_row}'] = "SUBTOTAL N"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = fill_subtotal
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'F{current_row}'].border = border_thin
    ws[f'F{current_row}'].fill = fill_subtotal
    current_row += 2  # Espacio

    # ===== SECCIÓN MATERIALES =====
    ws[f'A{current_row}'] = "MATERIALES"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].fill = fill_category
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 1

    # Headers MATERIALES (solo 5 columnas)
    headers_materiales = ['DESCRIPCION', 'UNIDAD', 'CANTIDAD', 'P. UNITARIO', 'COSTO']
    subheaders_materiales = ['', '', 'A', 'B', 'C = A x B']

    for col_idx, (header, subheader) in enumerate(zip(headers_materiales, subheaders_materiales), start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    for col_idx, subheader in enumerate(subheaders_materiales, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=subheader)
        cell.font = Font(italic=True, size=9)
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    # Filas de ejemplo para MATERIALES (3 filas)
    materiales_start_row = current_row
    for i in range(3):
        for col_idx in range(1, 6):  # Solo 5 columnas
            cell = ws.cell(row=current_row, column=col_idx, value="")
            cell.border = border_thin
        current_row += 1

    # Subtotal MATERIALES
    ws[f'A{current_row}'] = "SUBTOTAL O"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = fill_subtotal
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'E{current_row}'].border = border_thin
    ws[f'E{current_row}'].fill = fill_subtotal
    current_row += 2  # Espacio

    # ===== SECCIÓN TRANSPORTE =====
    ws[f'A{current_row}'] = "TRANSPORTE"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].fill = fill_category
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 1

    # Headers TRANSPORTE (similar a EQUIPOS)
    headers_transporte = ['DESCRIPCION', 'CANTIDAD', 'TARIFA', 'COSTO HORA', 'RENDIMIENTO', 'COSTO']
    subheaders_transporte = ['', 'A', 'B', 'C = A x B', 'R', 'D = C x R']

    for col_idx, (header, subheader) in enumerate(zip(headers_transporte, subheaders_transporte), start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    for col_idx, subheader in enumerate(subheaders_transporte, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=subheader)
        cell.font = Font(italic=True, size=9)
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    # Filas de ejemplo para TRANSPORTE (3 filas)
    transporte_start_row = current_row
    for i in range(3):
        for col_idx in range(1, 7):
            cell = ws.cell(row=current_row, column=col_idx, value="")
            cell.border = border_thin
        current_row += 1

    # Subtotal TRANSPORTE
    ws[f'A{current_row}'] = "SUBTOTAL P"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = fill_subtotal
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'F{current_row}'].border = border_thin
    ws[f'F{current_row}'].fill = fill_subtotal
    current_row += 2  # Espacio

    # ===== TOTALES =====
    ws[f'A{current_row}'] = "TOTAL COSTO DIRECTO"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'F{current_row}'].border = Border(
        left=Side(style='double'),
        right=Side(style='double'),
        top=Side(style='double'),
        bottom=Side(style='double')
    )

    # Guardar
    output_path = "/home/codevars/ETL_ET/data/templates/Template_APUS.xlsx"
    wb.save(output_path)
    print(f"Template creado exitosamente en: {output_path}")

    # Imprimir información sobre las filas
    print("\nEstructura del template:")
    print(f"- Info del rubro: filas 1-3")
    print(f"- EQUIPOS: inicio fila ~6, tabla ~{equipos_start_row}")
    print(f"- MANO DE OBRA: tabla ~{mano_start_row}")
    print(f"- MATERIALES: tabla ~{materiales_start_row}")
    print(f"- TRANSPORTE: tabla ~{transporte_start_row}")

if __name__ == "__main__":
    create_template()
