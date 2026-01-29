"""
Script de verificación para validar que el Excel generado tenga los datos correctos.
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ ERROR: openpyxl no está instalado")
    print("   Instala con: pip install openpyxl")
    sys.exit(1)


def verify_excel(file_path: Path):
    """Verifica que el Excel tenga los datos correctos."""

    if not file_path.exists():
        print(f"❌ ERROR: Archivo no encontrado: {file_path}")
        return False

    wb = load_workbook(file_path)

    print("=" * 70)
    print(f"VERIFICACIÓN: {file_path.name}")
    print("=" * 70)
    print()

    # Verificar hojas
    print(f"Hojas encontradas: {wb.sheetnames}")
    print()

    # Verificar cada hoja
    for sheet_name in wb.sheetnames:
        if sheet_name == "TEMPLATE":
            continue  # Skip la hoja template si existe

        print("-" * 70)
        print(f"HOJA: {sheet_name}")
        print("-" * 70)

        sheet = wb[sheet_name]

        # Leer datos del rubro
        codigo = sheet["B1"].value
        descripcion = sheet["B2"].value
        unidad = sheet["B3"].value

        print(f"CODIGO: {codigo}")
        print(f"DESCRIPCION: {descripcion}")
        print(f"UNIDAD: {unidad}")
        print()

        # Verificar sección EQUIPOS (fila 8+)
        print("SECCIÓN EQUIPOS (filas 8-10):")
        equipos_encontrados = []
        for row in range(8, 11):
            descripcion_equipo = sheet[f"A{row}"].value
            if descripcion_equipo:
                cantidad = sheet[f"B{row}"].value
                equipos_encontrados.append((descripcion_equipo, cantidad))
                print(f"  Fila {row}: {descripcion_equipo} (Cantidad: {cantidad})")

        if not equipos_encontrados:
            print("  (vacío)")
        print()

        # Verificar sección MANO DE OBRA (fila 16+)
        print("SECCIÓN MANO DE OBRA (filas 16-18):")
        mano_obra_encontrados = []
        for row in range(16, 19):
            descripcion_mo = sheet[f"A{row}"].value
            if descripcion_mo:
                cantidad = sheet[f"B{row}"].value
                mano_obra_encontrados.append((descripcion_mo, cantidad))
                print(f"  Fila {row}: {descripcion_mo} (Cantidad: {cantidad})")

        if not mano_obra_encontrados:
            print("  (vacío)")
        print()

        # Verificar sección MATERIALES (fila 24+)
        print("SECCIÓN MATERIALES (filas 24-26):")
        materiales_encontrados = []
        for row in range(24, 27):
            descripcion_mat = sheet[f"A{row}"].value
            if descripcion_mat:
                unidad_mat = sheet[f"B{row}"].value
                cantidad_mat = sheet[f"C{row}"].value
                materiales_encontrados.append((descripcion_mat, unidad_mat, cantidad_mat))
                print(f"  Fila {row}: {descripcion_mat} | {unidad_mat} | Cantidad: {cantidad_mat}")

        if not materiales_encontrados:
            print("  (vacío)")
        print()

        # Verificar sección TRANSPORTE (fila 32+)
        print("SECCIÓN TRANSPORTE (filas 32-34):")
        transporte_encontrados = []
        for row in range(32, 35):
            descripcion_trans = sheet[f"A{row}"].value
            if descripcion_trans:
                cantidad = sheet[f"B{row}"].value
                transporte_encontrados.append((descripcion_trans, cantidad))
                print(f"  Fila {row}: {descripcion_trans} (Cantidad: {cantidad})")

        if not transporte_encontrados:
            print("  (vacío)")
        print()

        # Validaciones específicas por rubro
        if sheet_name == "01.001.4.01":
            print("VALIDACIONES PARA RUBRO 01.001.4.01:")
            validaciones = []

            # Debe tener 3 materiales
            if len(materiales_encontrados) == 3:
                validaciones.append("✓ 3 materiales encontrados")
            else:
                validaciones.append(f"❌ Esperado 3 materiales, encontrado {len(materiales_encontrados)}")

            # Debe tener 2 equipos
            if len(equipos_encontrados) == 2:
                validaciones.append("✓ 2 equipos encontrados")
            else:
                validaciones.append(f"❌ Esperado 2 equipos, encontrado {len(equipos_encontrados)}")

            # Verificar nombres específicos
            mat_nombres = [m[0] for m in materiales_encontrados]
            if "ESTACAS" in mat_nombres:
                validaciones.append("✓ Material 'ESTACAS' encontrado")
            else:
                validaciones.append("❌ Material 'ESTACAS' NO encontrado")

            equipo_nombres = [e[0] for e in equipos_encontrados]
            if "HERRAMIENTA MENOR" in equipo_nombres:
                validaciones.append("✓ Equipo 'HERRAMIENTA MENOR' encontrado")
            else:
                validaciones.append("❌ Equipo 'HERRAMIENTA MENOR' NO encontrado")

            for val in validaciones:
                print(val)
            print()

        elif sheet_name == "01.002.4.01":
            print("VALIDACIONES PARA RUBRO 01.002.4.01:")
            validaciones = []

            # No debe tener materiales
            if len(materiales_encontrados) == 0:
                validaciones.append("✓ Sin materiales (No aplica)")
            else:
                validaciones.append(f"❌ Esperado 0 materiales, encontrado {len(materiales_encontrados)}")

            # Debe tener 1 equipo
            if len(equipos_encontrados) == 1:
                validaciones.append("✓ 1 equipo encontrado")
            else:
                validaciones.append(f"❌ Esperado 1 equipo, encontrado {len(equipos_encontrados)}")

            for val in validaciones:
                print(val)
            print()

    print("=" * 70)
    print("VERIFICACIÓN COMPLETADA")
    print("=" * 70)
    print()

    wb.close()
    return True


def main():
    """Ejecuta la verificación."""
    file_path = Path("data/output/TEST_APUS_TEMPLATE.xlsx")

    if not file_path.exists():
        print(f"❌ ERROR: Archivo no encontrado: {file_path}")
        print("   Ejecuta primero: python scripts/test_template_export.py")
        return 1

    try:
        verify_excel(file_path)
        return 0
    except Exception as e:
        print()
        print("=" * 70)
        print("❌ ERROR DURANTE LA VERIFICACIÓN")
        print("=" * 70)
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
