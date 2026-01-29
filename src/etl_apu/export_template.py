"""
Exportador de APUs usando template Excel.

Genera un archivo Excel con:
- 1 hoja por rubro (nombre = CODIGO del rubro)
- Datos mapeados según configuración de template
- Inserción dinámica de filas para recursos
- Preservación de headers y formato del template

Soporta templates .xls (convierte a .xlsx) y .xlsx
"""

from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
import logging
from copy import copy

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Fill, Border, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

logger = logging.getLogger(__name__)


@dataclass
class TemplateMapping:
    """
    Configuración de mapeo de celdas del template.

    Define dónde escribir cada dato en la hoja del template.
    Ajustar según el template real.
    """
    # Datos del rubro
    codigo_cell: str = "B1"
    descripcion_cell: str = "B2"
    unidad_cell: str = "B3"

    # Sección EQUIPOS (6 columnas: DESCRIPCION, CANTIDAD, TARIFA, COSTO HORA, RENDIMIENTO, COSTO)
    equipos_start_row: int = 8
    equipos_cols: Tuple[str, ...] = ("A", "B", "C", "D", "E", "F")
    equipos_subtotal_row: int = 11

    # Sección MANO DE OBRA (6 columnas: DESCRIPCION, CANTIDAD, JORNAL/HR, COSTO HORA, RENDIMIENTO, COSTO)
    mano_obra_start_row: int = 16
    mano_obra_cols: Tuple[str, ...] = ("A", "B", "C", "D", "E", "F")
    mano_obra_subtotal_row: int = 19

    # Sección MATERIALES (5 columnas: DESCRIPCION, UNIDAD, CANTIDAD, P. UNITARIO, COSTO)
    materiales_start_row: int = 24
    materiales_cols: Tuple[str, ...] = ("A", "B", "C", "D", "E")
    materiales_subtotal_row: int = 27

    # Sección TRANSPORTE (6 columnas: DESCRIPCION, CANTIDAD, TARIFA, COSTO HORA, RENDIMIENTO, COSTO)
    transporte_start_row: int = 32
    transporte_cols: Tuple[str, ...] = ("A", "B", "C", "D", "E", "F")
    transporte_subtotal_row: int = 35

    # Filas por defecto por categoría (3 filas de recursos)
    default_rows_per_category: int = 3

    # Observaciones
    observaciones_cell: Optional[str] = None


@dataclass
class ExportStats:
    """Estadísticas del export."""
    total_rubros: int = 0
    hojas_creadas: int = 0
    duplicados_detectados: int = 0
    filas_insertadas: int = 0
    recursos_exportados: int = 0
    warnings: List[str] = None

    def __post_init__(self):
        if self.warnings is None:
            self.warnings = []


class TemplateExporter:
    """
    Exportador de APUs basado en template Excel.

    Genera 1 hoja por rubro usando un template como base.
    """

    def __init__(
        self,
        template_path: Path,
        mapping: Optional[TemplateMapping] = None
    ):
        """
        Inicializa el exportador.

        Args:
            template_path: Ruta al template Excel (.xls o .xlsx)
            mapping: Configuración de mapeo de celdas

        Raises:
            RuntimeError: Si openpyxl no está instalado
            FileNotFoundError: Si el template no existe
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl no instalado. "
                "Instalar con: pip install openpyxl"
            )

        self.template_path = template_path
        self.mapping = mapping or TemplateMapping()
        self.sheet_names_used = set()

        if not template_path.exists():
            raise FileNotFoundError(f"Template no encontrado: {template_path}")

        # Convertir .xls a .xlsx si es necesario
        if template_path.suffix == ".xls":
            logger.info("Template es .xls, convirtiendo a .xlsx...")
            self.template_path = self._convert_xls_to_xlsx(template_path)

        # Cargar template
        logger.info(f"Cargando template: {self.template_path}")
        self.template_wb = load_workbook(self.template_path)
        self.template_sheet_name = self.template_wb.sheetnames[0]

        logger.info(f"Template cargado: hoja base '{self.template_sheet_name}'")

    def _convert_xls_to_xlsx(self, xls_path: Path) -> Path:
        """
        Convierte .xls a .xlsx para poder usar openpyxl.

        Args:
            xls_path: Ruta al archivo .xls

        Returns:
            Ruta al archivo .xlsx generado

        Raises:
            RuntimeError: Si xlrd no está instalado o falla la conversión
        """
        if not XLRD_AVAILABLE:
            raise RuntimeError(
                "xlrd no instalado. Necesario para leer .xls\n"
                "Instalar con: pip install xlrd"
            )

        # TODO: Implementar conversión real con xlrd → openpyxl
        # Por ahora, asumir que el usuario proporciona .xlsx

        raise RuntimeError(
            f"Template es .xls: {xls_path}\n"
            "Por favor, convierte manualmente a .xlsx o usa un template .xlsx"
        )

    def export_apus(
        self,
        output_path: Path,
        rubros_data: List[Dict],
        recursos_por_rubro: Dict[str, List[Dict]]
    ) -> ExportStats:
        """
        Exporta APUs a Excel usando el template.

        Args:
            output_path: Ruta donde guardar el Excel de salida
            rubros_data: Lista de dicts con datos de rubros
                [{"codigo": "01.001.4.01", "descripcion": "...", "unidad": "m2"}, ...]
            recursos_por_rubro: Dict {codigo_rubro: [recursos]}
                {"01.001.4.01": [{"categoria": "MATERIALES", "nombre": "Estacas", ...}]}

        Returns:
            ExportStats con estadísticas del export
        """
        logger.info(f"Iniciando export tipo APU a {output_path}")
        logger.info(f"Rubros a exportar: {len(rubros_data)}")

        stats = ExportStats(total_rubros=len(rubros_data))

        # Crear workbook de salida
        output_wb = Workbook()
        output_wb.remove(output_wb.active)  # Remover hoja default vacía

        # Procesar cada rubro
        for rubro in rubros_data:
            codigo = rubro["codigo"]
            recursos = recursos_por_rubro.get(codigo, [])

            # Crear hoja para este rubro
            sheet_name = self._get_unique_sheet_name(codigo, stats)

            logger.info(f"Creando hoja '{sheet_name}' para rubro {codigo}")

            # Clonar template
            new_sheet = self._clone_template_sheet(output_wb, sheet_name)

            # Escribir datos del rubro
            self._write_rubro_data(new_sheet, rubro)

            # Escribir recursos
            filas_insertadas = self._write_recursos(new_sheet, recursos)

            stats.hojas_creadas += 1
            stats.filas_insertadas += filas_insertadas
            stats.recursos_exportados += len(recursos)

        # Guardar archivo
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_wb.save(output_path)

        logger.info(f"✅ Excel generado: {output_path}")
        logger.info(f"   Hojas creadas: {stats.hojas_creadas}")
        logger.info(f"   Recursos exportados: {stats.recursos_exportados}")
        logger.info(f"   Filas insertadas: {stats.filas_insertadas}")

        if stats.duplicados_detectados > 0:
            logger.warning(f"   ⚠️  Duplicados detectados: {stats.duplicados_detectados}")

        return stats

    def _get_unique_sheet_name(self, codigo: str, stats: ExportStats) -> str:
        """
        Genera nombre único de hoja manejando duplicados.

        Args:
            codigo: Código del rubro
            stats: Stats para trackear duplicados

        Returns:
            Nombre de hoja único
        """
        # Sanitizar nombre (Excel no permite algunos caracteres)
        base_name = codigo.replace("/", "_").replace("\\", "_")[:31]

        # Si es único, usar directamente
        if base_name not in self.sheet_names_used:
            self.sheet_names_used.add(base_name)
            return base_name

        # Duplicado detectado
        stats.duplicados_detectados += 1
        stats.warnings.append(
            f"Código duplicado: {codigo} (generando nombre con sufijo)"
        )

        logger.warning(f"⚠️  Código duplicado: {codigo}")

        # Agregar sufijo incremental
        counter = 2
        while True:
            unique_name = f"{base_name[:27]} ({counter})"
            if unique_name not in self.sheet_names_used:
                self.sheet_names_used.add(unique_name)
                return unique_name
            counter += 1

    def _clone_template_sheet(self, target_wb: Workbook, new_name: str):
        """
        Clona la hoja del template al workbook de salida.

        Args:
            target_wb: Workbook destino
            new_name: Nombre de la nueva hoja

        Returns:
            Nueva hoja clonada
        """
        # Obtener hoja template
        template_sheet = self.template_wb[self.template_sheet_name]

        # Crear nueva hoja
        new_sheet = target_wb.create_sheet(title=new_name)

        # Copiar dimensiones de columnas
        for col_letter in template_sheet.column_dimensions:
            if col_letter in template_sheet.column_dimensions:
                new_sheet.column_dimensions[col_letter].width = \
                    template_sheet.column_dimensions[col_letter].width

        # Copiar dimensiones de filas
        for row_num in template_sheet.row_dimensions:
            if row_num in template_sheet.row_dimensions:
                new_sheet.row_dimensions[row_num].height = \
                    template_sheet.row_dimensions[row_num].height

        # Copiar celdas con valores y estilos
        for row in template_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value

                # Copiar estilos
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copiar merged cells
        for merged_range in template_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))

        return new_sheet

    def _write_rubro_data(self, sheet, rubro: Dict):
        """
        Escribe datos del rubro en las celdas configuradas.

        Args:
            sheet: Hoja de Excel
            rubro: Dict con datos del rubro
        """
        # Escribir código
        if self.mapping.codigo_cell:
            sheet[self.mapping.codigo_cell] = rubro.get("codigo", "")

        # Escribir descripción
        if self.mapping.descripcion_cell:
            sheet[self.mapping.descripcion_cell] = rubro.get("descripcion", "")

        # Escribir unidad
        if self.mapping.unidad_cell:
            sheet[self.mapping.unidad_cell] = rubro.get("unidad", "")

    def _write_recursos(self, sheet, recursos: List[Dict]) -> int:
        """
        Escribe recursos en las secciones del template según categoría.

        Cada categoría se escribe en su sección específica del template.
        Inserta filas dinámicamente si hay más recursos que filas disponibles.

        Args:
            sheet: Hoja de Excel
            recursos: Lista de recursos

        Returns:
            Número de filas insertadas
        """
        if not recursos:
            return 0

        filas_insertadas = 0

        # Agrupar por categoría
        recursos_por_categoria = self._agrupar_recursos_por_categoria(recursos)

        # Mapeo de categorías a configuración
        categoria_config = {
            "EQUIPO": {
                "start_row": self.mapping.equipos_start_row,
                "cols": self.mapping.equipos_cols,
                "subtotal_row": self.mapping.equipos_subtotal_row,
            },
            "MANO_OBRA": {
                "start_row": self.mapping.mano_obra_start_row,
                "cols": self.mapping.mano_obra_cols,
                "subtotal_row": self.mapping.mano_obra_subtotal_row,
            },
            "MATERIALES": {
                "start_row": self.mapping.materiales_start_row,
                "cols": self.mapping.materiales_cols,
                "subtotal_row": self.mapping.materiales_subtotal_row,
            },
            "TRANSPORTE": {
                "start_row": self.mapping.transporte_start_row,
                "cols": self.mapping.transporte_cols,
                "subtotal_row": self.mapping.transporte_subtotal_row,
            },
        }

        # Procesar cada categoría
        for categoria, config in categoria_config.items():
            if categoria not in recursos_por_categoria:
                continue

            recursos_cat = recursos_por_categoria[categoria]

            # Escribir recursos de esta categoría
            inserted = self._write_categoria_recursos(
                sheet=sheet,
                recursos=recursos_cat,
                categoria=categoria,
                start_row=config["start_row"],
                cols=config["cols"],
                subtotal_row=config["subtotal_row"]
            )

            filas_insertadas += inserted

        return filas_insertadas

    def _write_categoria_recursos(
        self,
        sheet,
        recursos: List[Dict],
        categoria: str,
        start_row: int,
        cols: Tuple[str, ...],
        subtotal_row: int
    ) -> int:
        """
        Escribe recursos de una categoría específica en su sección.

        Args:
            sheet: Hoja de Excel
            recursos: Lista de recursos de la categoría
            categoria: Nombre de la categoría
            start_row: Fila inicial de datos
            cols: Tupla de letras de columnas
            subtotal_row: Fila del subtotal

        Returns:
            Número de filas insertadas
        """
        if not recursos:
            return 0

        filas_insertadas = 0
        filas_disponibles = subtotal_row - start_row

        # Si necesitamos más filas, insertar antes del subtotal
        if len(recursos) > filas_disponibles:
            filas_a_insertar = len(recursos) - filas_disponibles

            # Insertar filas antes del subtotal
            sheet.insert_rows(subtotal_row, filas_a_insertar)

            # Copiar formato de la fila base (última fila de recursos)
            fila_base = subtotal_row - 1
            for i in range(filas_a_insertar):
                nueva_fila = subtotal_row + i
                for col in cols:
                    cell_base = sheet[f"{col}{fila_base}"]
                    cell_nueva = sheet[f"{col}{nueva_fila}"]

                    # Copiar estilos
                    if cell_base.has_style:
                        cell_nueva.font = copy(cell_base.font)
                        cell_nueva.border = copy(cell_base.border)
                        cell_nueva.fill = copy(cell_base.fill)
                        cell_nueva.number_format = copy(cell_base.number_format)
                        cell_nueva.alignment = copy(cell_base.alignment)

            filas_insertadas = filas_a_insertar

        # Escribir recursos
        for idx, recurso in enumerate(recursos):
            current_row = start_row + idx

            # Columna A: DESCRIPCION (siempre)
            sheet[f"{cols[0]}{current_row}"] = recurso.get("nombre", "")

            # Las siguientes columnas dependen de la categoría
            if categoria == "MATERIALES":
                # MATERIALES: DESCRIPCION | UNIDAD | CANTIDAD | P. UNITARIO | COSTO
                if len(cols) >= 2:
                    sheet[f"{cols[1]}{current_row}"] = recurso.get("unidad", "")
                if len(cols) >= 3:
                    sheet[f"{cols[2]}{current_row}"] = recurso.get("cantidad", "")
                # P. UNITARIO y COSTO pueden dejarse en blanco o calcularse
            else:
                # EQUIPOS/MANO_OBRA/TRANSPORTE: DESCRIPCION | CANTIDAD | TARIFA | COSTO HORA | RENDIMIENTO | COSTO
                if len(cols) >= 2:
                    sheet[f"{cols[1]}{current_row}"] = recurso.get("cantidad", "")
                # Las demás columnas (TARIFA, COSTO HORA, etc.) pueden dejarse para cálculo manual

        return filas_insertadas

    def _agrupar_recursos_por_categoria(
        self,
        recursos: List[Dict]
    ) -> Dict[str, List[Dict]]:
        """
        Agrupa recursos por categoría.

        Args:
            recursos: Lista de recursos

        Returns:
            Dict {categoria: [recursos]}
        """
        agrupados = {}

        for recurso in recursos:
            categoria = recurso.get("categoria", "DESCONOCIDO")
            if categoria not in agrupados:
                agrupados[categoria] = []
            agrupados[categoria].append(recurso)

        return agrupados


# ═══════════════════════════════════════════════════════════════════════════
# FUNCIONES DE CONVENIENCIA
# ═══════════════════════════════════════════════════════════════════════════

def export_apus_from_template(
    template_path: Path,
    output_path: Path,
    rubros_data: List[Dict],
    recursos_por_rubro: Dict[str, List[Dict]],
    mapping: Optional[TemplateMapping] = None
) -> ExportStats:
    """
    Función de conveniencia para exportar APUs desde template.

    Args:
        template_path: Ruta al template Excel
        output_path: Ruta de salida
        rubros_data: Lista de rubros
        recursos_por_rubro: Dict de recursos por código de rubro
        mapping: Configuración de mapeo (opcional)

    Returns:
        ExportStats con estadísticas
    """
    exporter = TemplateExporter(template_path, mapping)
    return exporter.export_apus(output_path, rubros_data, recursos_por_rubro)
