"""
Módulo de exportación a Excel.

Responsabilidades:
- Generar archivo Excel con múltiples hojas
- Estructurar Rubros, Recursos, Warnings en tablas
- Aplicar formato profesional (anchos, colores, etc.)
- Validar datos antes de exportar
"""

from pathlib import Path
from typing import List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from src.models.schemas import (
    Rubro, Recurso, ParseWarning,
    PipelineResult, DocumentMetadata
)
from src.utils.logger import get_logger

logger = get_logger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# CONVERSIÓN A DATAFRAMES
# ═══════════════════════════════════════════════════════════════════════════

def rubros_to_dataframe(rubros: List[Rubro]) -> pd.DataFrame:
    """
    Convierte lista de Rubros a DataFrame.

    Args:
        rubros: Lista de objetos Rubro

    Returns:
        DataFrame con columnas normalizadas

    Columns:
        - rubro_id: ID único
        - codigo: Código del rubro
        - descripcion: Descripción
        - unidad: Unidad de medida
        - pages: Páginas de origen (separadas por coma)
        - confidence: Score de confianza
        - metodo_constructivo: Método (si existe)
    """
    if not rubros:
        return pd.DataFrame(columns=[
            'rubro_id', 'codigo', 'descripcion', 'unidad',
            'pages', 'confidence', 'metodo_constructivo'
        ])

    data = []
    for rubro in rubros:
        data.append({
            'rubro_id': rubro.rubro_id,
            'codigo': rubro.codigo,
            'descripcion': rubro.descripcion,
            'unidad': rubro.unidad,
            'pages': ', '.join(map(str, rubro.source_pages)),
            'confidence': round(rubro.confidence, 2),
            'metodo_constructivo': rubro.metodo_constructivo or ''
        })

    df = pd.DataFrame(data)

    # Ordenar por código de rubro
    df = df.sort_values('codigo').reset_index(drop=True)

    return df


def recursos_to_dataframe(recursos: List[Recurso]) -> pd.DataFrame:
    """
    Convierte lista de Recursos a DataFrame.

    Args:
        recursos: Lista de objetos Recurso

    Returns:
        DataFrame con columnas normalizadas

    Columns:
        - recurso_id: ID único
        - rubro_id: ID del rubro padre
        - tipo: MATERIAL | EQUIPO | DESCONOCIDO
        - nombre: Descripción del recurso
        - unidad: Unidad (si existe)
        - cantidad: Cantidad (si existe)
        - confidence: Score de confianza
    """
    if not recursos:
        return pd.DataFrame(columns=[
            'recurso_id', 'rubro_id', 'tipo', 'nombre',
            'unidad', 'cantidad', 'confidence'
        ])

    data = []
    for recurso in recursos:
        data.append({
            'recurso_id': recurso.recurso_id,
            'rubro_id': recurso.rubro_id,
            'tipo': recurso.tipo,
            'nombre': recurso.nombre,
            'unidad': recurso.unidad or '',
            'cantidad': recurso.cantidad if recurso.cantidad else '',
            'confidence': round(recurso.confidence, 2)
        })

    df = pd.DataFrame(data)

    # Ordenar por rubro_id y recurso_id
    df = df.sort_values(['rubro_id', 'recurso_id']).reset_index(drop=True)

    return df


def warnings_to_dataframe(warnings: List[ParseWarning]) -> pd.DataFrame:
    """
    Convierte lista de Warnings a DataFrame.

    Args:
        warnings: Lista de objetos ParseWarning

    Returns:
        DataFrame con columnas normalizadas

    Columns:
        - warning_id: ID único
        - rubro_id: ID del rubro (si aplica)
        - page: Número de página
        - kind: Tipo de warning
        - severity: LOW | MEDIUM | HIGH
        - message: Mensaje descriptivo
        - snippet: Fragmento de texto (primeros 100 chars)
    """
    if not warnings:
        return pd.DataFrame(columns=[
            'warning_id', 'rubro_id', 'page', 'kind',
            'severity', 'message', 'snippet'
        ])

    data = []
    for warning in warnings:
        data.append({
            'warning_id': warning.warning_id,
            'rubro_id': warning.rubro_id or '',
            'page': warning.page or '',
            'kind': warning.kind,
            'severity': warning.severity,
            'message': warning.message,
            'snippet': (warning.snippet[:100] + '...') if warning.snippet and len(warning.snippet) > 100 else (warning.snippet or '')
        })

    df = pd.DataFrame(data)

    # Ordenar por severidad (HIGH → MEDIUM → LOW) y página
    severity_order = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
    df['severity_order'] = df['severity'].map(severity_order)
    df = df.sort_values(['severity_order', 'page']).drop('severity_order', axis=1)
    df = df.reset_index(drop=True)

    return df


def metadata_to_dataframe(metadata: DocumentMetadata) -> pd.DataFrame:
    """
    Convierte metadatos del documento a DataFrame (para hoja de resumen).

    Args:
        metadata: Objeto DocumentMetadata

    Returns:
        DataFrame con 2 columnas: [Campo, Valor]
    """
    data = {
        'Campo': [
            'Archivo',
            'Total Páginas',
            'Tipo de Documento',
            'Páginas con OCR',
            'Total Rubros',
            'Total Recursos',
            'Total Warnings',
            'Fecha de Procesamiento'
        ],
        'Valor': [
            metadata.filename,
            metadata.total_pages,
            metadata.tipo_documento,
            ', '.join(map(str, metadata.pages_with_ocr)) if metadata.pages_with_ocr else 'Ninguna',
            metadata.total_rubros,
            metadata.total_recursos,
            metadata.total_warnings,
            metadata.processing_date.strftime('%Y-%m-%d %H:%M:%S')
        ]
    }

    return pd.DataFrame(data)


# ═══════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN A EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def export_to_excel(
    result: PipelineResult,
    output_path: Path,
    apply_formatting: bool = True
) -> None:
    """
    Exporta resultados del pipeline a un archivo Excel con múltiples hojas.

    Hojas creadas:
    1. Resumen: Metadatos del documento y estadísticas
    2. Rubros: Tabla de rubros
    3. Recursos: Tabla de recursos
    4. Warnings: Tabla de warnings/errores
    5. Relaciones: Tabla de relaciones Rubro → Recursos

    Args:
        result: Objeto PipelineResult con todos los datos
        output_path: Ruta donde guardar el Excel
        apply_formatting: Si True, aplica formato (colores, anchos, etc.)

    Raises:
        ValueError: Si result está vacío
        IOError: Si no se puede escribir el archivo

    Example:
        >>> result = PipelineResult(...)
        >>> output_path = Path("data/output/resultado.xlsx")
        >>> export_to_excel(result, output_path)
    """
    logger.info(f"Exportando resultados a Excel: {output_path}")

    # Validar que hay datos
    if not result.rubros and not result.recursos:
        logger.warning("No hay rubros ni recursos para exportar")

    # Crear DataFrames
    df_metadata = metadata_to_dataframe(result.metadata)
    df_rubros = rubros_to_dataframe(result.rubros)
    df_recursos = recursos_to_dataframe(result.recursos)
    df_warnings = warnings_to_dataframe(result.warnings)
    df_relaciones = crear_tabla_relaciones(result.rubros, result.recursos)

    # Crear archivo Excel con múltiples hojas
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_metadata.to_excel(writer, sheet_name='Resumen', index=False)
            df_rubros.to_excel(writer, sheet_name='Rubros', index=False)
            df_recursos.to_excel(writer, sheet_name='Recursos', index=False)
            df_relaciones.to_excel(writer, sheet_name='Relaciones', index=False)
            df_warnings.to_excel(writer, sheet_name='Warnings', index=False)

        # Aplicar formato si está habilitado
        if apply_formatting:
            aplicar_formato_excel(output_path)

        logger.info(f"Excel exportado exitosamente: {output_path}")

    except Exception as e:
        logger.error(f"Error al exportar Excel: {e}")
        raise IOError(f"No se pudo guardar el archivo Excel: {e}")


def crear_tabla_relaciones(
    rubros: List[Rubro],
    recursos: List[Recurso]
) -> pd.DataFrame:
    """
    Crea tabla de relaciones Rubro → Recursos (para análisis).

    Args:
        rubros: Lista de rubros
        recursos: Lista de recursos

    Returns:
        DataFrame con columnas: [rubro_codigo, rubro_desc, recurso_tipo, recurso_nombre, cantidad]
    """
    if not rubros or not recursos:
        return pd.DataFrame(columns=[
            'rubro_codigo', 'rubro_descripcion', 'recurso_tipo',
            'recurso_nombre', 'cantidad', 'unidad'
        ])

    # Crear mapa rubro_id → rubro
    rubros_map = {r.rubro_id: r for r in rubros}

    data = []
    for recurso in recursos:
        rubro = rubros_map.get(recurso.rubro_id)
        if rubro:
            data.append({
                'rubro_codigo': rubro.codigo,
                'rubro_descripcion': rubro.descripcion,
                'recurso_tipo': recurso.tipo,
                'recurso_nombre': recurso.nombre,
                'cantidad': recurso.cantidad if recurso.cantidad else '',
                'unidad': recurso.unidad or ''
            })

    df = pd.DataFrame(data)
    df = df.sort_values(['rubro_codigo', 'recurso_tipo']).reset_index(drop=True)

    return df


# ═══════════════════════════════════════════════════════════════════════════
# FORMATO Y ESTILO
# ═══════════════════════════════════════════════════════════════════════════

def aplicar_formato_excel(file_path: Path) -> None:
    """
    Aplica formato profesional al archivo Excel.

    - Ajusta anchos de columnas
    - Aplica colores a encabezados
    - Colorea warnings por severidad
    - Congela paneles

    Args:
        file_path: Ruta al archivo Excel ya creado
    """
    logger.debug(f"Aplicando formato a {file_path}")

    try:
        wb = load_workbook(file_path)

        # Formato para cada hoja
        hojas_config = {
            'Resumen': {'freeze': 'A2', 'header_color': '4472C4'},
            'Rubros': {'freeze': 'A2', 'header_color': '70AD47'},
            'Recursos': {'freeze': 'A2', 'header_color': 'FFC000'},
            'Relaciones': {'freeze': 'A2', 'header_color': '5B9BD5'},
            'Warnings': {'freeze': 'A2', 'header_color': 'C00000'}
        }

        for sheet_name, config in hojas_config.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Formato de encabezados
                header_fill = PatternFill(start_color=config['header_color'],
                                         end_color=config['header_color'],
                                         fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Ajustar anchos de columna
                ajustar_anchos_columna(ws)

                # Congelar paneles
                ws.freeze_panes = config['freeze']

        # Colorear warnings por severidad
        if 'Warnings' in wb.sheetnames:
            colorear_warnings(wb['Warnings'])

        wb.save(file_path)
        logger.debug("Formato aplicado exitosamente")

    except Exception as e:
        logger.warning(f"No se pudo aplicar formato: {e}")


def ajustar_anchos_columna(worksheet) -> None:
    """
    Ajusta anchos de columnas automáticamente basado en contenido.

    Args:
        worksheet: Worksheet de openpyxl
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
        worksheet.column_dimensions[column_letter].width = adjusted_width


def colorear_warnings(worksheet) -> None:
    """
    Colorea filas de warnings según severidad.

    - HIGH: Rojo claro
    - MEDIUM: Amarillo
    - LOW: Gris claro

    Args:
        worksheet: Worksheet de openpyxl (hoja "Warnings")
    """
    # Encontrar índice de columna 'severity'
    headers = [cell.value for cell in worksheet[1]]
    if 'severity' not in headers:
        return

    severity_col_idx = headers.index('severity') + 1  # 1-indexed

    colors = {
        'HIGH': 'FFCCCC',    # Rojo claro
        'MEDIUM': 'FFFFCC',  # Amarillo claro
        'LOW': 'E7E6E6'      # Gris claro
    }

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        severity = row[severity_col_idx - 1].value

        if severity in colors:
            fill = PatternFill(start_color=colors[severity],
                             end_color=colors[severity],
                             fill_type='solid')

            for cell in row:
                cell.fill = fill


# ═══════════════════════════════════════════════════════════════════════════
# VALIDACIÓN PRE-EXPORT
# ═══════════════════════════════════════════════════════════════════════════

def validar_antes_de_exportar(result: PipelineResult) -> List[str]:
    """
    Valida datos antes de exportar (quality checks).

    Args:
        result: PipelineResult a validar

    Returns:
        Lista de errores encontrados (vacía si todo OK)

    Example:
        >>> errores = validar_antes_de_exportar(result)
        >>> if errores:
        >>>     print("\\n".join(errores))
    """
    errores = []

    # Check 1: Al menos un rubro
    if not result.rubros:
        errores.append("No hay rubros para exportar")

    # Check 2: Recursos huérfanos (sin rubro asociado)
    rubros_ids = {r.rubro_id for r in result.rubros}
    for recurso in result.recursos:
        if recurso.rubro_id not in rubros_ids:
            errores.append(
                f"Recurso {recurso.recurso_id} referencia rubro inexistente {recurso.rubro_id}"
            )

    # Check 3: Warnings huérfanos
    for warning in result.warnings:
        if warning.rubro_id and warning.rubro_id not in rubros_ids:
            errores.append(
                f"Warning {warning.warning_id} referencia rubro inexistente {warning.rubro_id}"
            )

    return errores
